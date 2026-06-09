"""Google Sheets data access layer."""
import io
import re
from typing import Any, Optional, Tuple, List

import gspread
import numpy as np
import openpyxl
import pandas as pd
import requests
from oauth2client.service_account import ServiceAccountCredentials
import streamlit as st

from ..config import Config


class GoogleSheetsClient:
    """Client for accessing Google Sheets data."""

    def __init__(self):
        """Initialize the Google Sheets client."""
        self._client = None

    @staticmethod
    @st.cache_resource
    def _get_cached_client() -> gspread.Client:
        """Create and cache the authenticated Google Sheets client."""
        scope = Config.get_google_api_scope()
        creds_dict = Config.get_gcp_credentials()
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        return gspread.authorize(creds)

    def _get_client(self) -> gspread.Client:
        """Get authenticated Google Sheets client.

        Returns:
            Authenticated gspread client.
        """
        if self._client is None:
            self._client = self._get_cached_client()
        return self._client

    @staticmethod
    @st.cache_data(ttl=Config.CACHE_TTL, show_spinner=False)
    def _download_public_xlsx(url: str) -> bytes:
        """Download a link-viewable spreadsheet as an .xlsx workbook (read-only).

        Used only when no service account is configured (local/dev); production
        uses the authenticated client. Never writes to the sheet.
        """
        match = re.search(r"/d/([A-Za-z0-9_-]+)", url)
        if not match:
            raise ValueError(f"Could not extract spreadsheet id from URL: {url}")
        xlsx_url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
        response = requests.get(xlsx_url, timeout=Config.REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.content

    @staticmethod
    def _cell_to_display(value: Any, number_format: Optional[str]) -> str:
        """Render a cell the way gspread's get_all_values() would (displayed text).

        Critically, percent-formatted cells are stored as fractions in the xlsx
        (0.121) but display as '12.1%'. Scaling them back keeps the existing
        parsers seeing the same magnitude they get from gspread in production.
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, (int, float)) and number_format and number_format.rstrip().endswith("%"):
            return f"{value * 100:g}%"
        return str(value)

    @staticmethod
    def _slice_a1_range(grid: List[List[str]], cell_range: str) -> List[List[str]]:
        """Slice an A1-notation block (e.g. 'O1:W22') out of a full grid."""
        match = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", cell_range)
        if not match:
            return grid

        def column_index(letters: str) -> int:
            index = 0
            for char in letters:
                index = index * 26 + (ord(char) - ord("A") + 1)
            return index - 1

        col_start, col_end = column_index(match.group(1)), column_index(match.group(3))
        row_start, row_end = int(match.group(2)) - 1, int(match.group(4)) - 1
        return [row[col_start:col_end + 1] for row in grid[row_start:row_end + 1]]

    @staticmethod
    def _read_public_grid(url: str, worksheet_name: str, cell_range: Optional[str] = None) -> List[List[str]]:
        """Mirror gspread reads (get_all_values / get(range)) from a public sheet."""
        workbook = openpyxl.load_workbook(
            io.BytesIO(GoogleSheetsClient._download_public_xlsx(url)),
            read_only=True,
            data_only=True,
        )
        try:
            if worksheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet {worksheet_name!r} not found. Available: {workbook.sheetnames}"
                )
            worksheet = workbook[worksheet_name]
            grid = [
                [GoogleSheetsClient._cell_to_display(cell.value, cell.number_format) for cell in row]
                for row in worksheet.iter_rows()
            ]
        finally:
            workbook.close()
        # gspread's get_all_values() stops at the last non-empty row; openpyxl
        # can include trailing blank rows from the sheet's used range.
        while grid and not any(cell.strip() for cell in grid[-1]):
            grid.pop()
        return GoogleSheetsClient._slice_a1_range(grid, cell_range) if cell_range else grid

    @staticmethod
    @st.cache_data(ttl=Config.CACHE_TTL)
    def _fetch_sheet_values(url: str, worksheet_name: str) -> List[List[str]]:
        """Fetch all values from a worksheet."""
        if not Config.has_service_account():
            return GoogleSheetsClient._read_public_grid(url, worksheet_name)
        client = GoogleSheetsClient._get_cached_client()
        return client.open_by_url(url).worksheet(worksheet_name).get_all_values()

    @staticmethod
    @st.cache_data(ttl=Config.CACHE_TTL)
    def _fetch_sheet_range(url: str, worksheet_name: str, cell_range: str) -> List[List[str]]:
        """Fetch a specific range from a worksheet."""
        if not Config.has_service_account():
            return GoogleSheetsClient._read_public_grid(url, worksheet_name, cell_range)
        client = GoogleSheetsClient._get_cached_client()
        return client.open_by_url(url).worksheet(worksheet_name).get(cell_range)

    @st.cache_data(ttl=Config.CACHE_TTL)
    def fetch_battery_data(_self, username_filter: Optional[str] = None) -> Tuple[pd.DataFrame, Optional[str]]:
        """Fetch battery data from Google Sheets.

        Args:
            username_filter: Optional username to filter results.

        Returns:
            Tuple of (DataFrame with battery data, battery pack column name).
        """
        try:
            url = Config.get_spreadsheet_url()
            data = _self._fetch_sheet_values(url, "Database")

            if not data:
                return pd.DataFrame(), None

            header = data[0]
            if 'Username' not in header:
                st.error("The 'Username' column is missing from the Google Sheets data.")
                return pd.DataFrame(), None

            filtered_columns = [
                (index, column.strip())
                for index, column in enumerate(header)
                if column and not column.startswith('_') and column not in Config.EXCLUDE_COLUMNS
            ]

            filtered_header = [column for _, column in filtered_columns]
            unique_header = _self._make_unique_headers(filtered_header)
            filtered_rows = [
                [row[index] if index < len(row) else "" for index, _ in filtered_columns]
                for row in data[1:]
            ]

            df = pd.DataFrame(filtered_rows, columns=unique_header)
            if df.empty:
                return df, None

            df = _self._process_dataframe(df)
            battery_pack_col = _self._find_battery_pack_column(df)

            if username_filter and 'Username' in df.columns:
                df = df[df['Username'].str.contains(username_filter, case=False, na=False)]

            return df.reset_index(drop=True), battery_pack_col

        except Exception as exc:
            st.error(f"Error fetching data from Google Sheets: {exc}")
            return pd.DataFrame(), None

    @st.cache_data(ttl=Config.CACHE_TTL)
    def fetch_battery_info(_self) -> pd.DataFrame:
        """Fetch battery pack information from Backend worksheet.

        Returns:
            DataFrame with battery pack specifications.
        """
        try:
            url = Config.get_spreadsheet_url()
            data = _self._fetch_sheet_range(url, "Backend", "O1:W22")
            if not data:
                return pd.DataFrame()

            header = data[0]
            rows = data[1:]
            battery_info = pd.DataFrame(rows, columns=header)
            battery_info = battery_info.replace("", np.nan).dropna(axis=1, how='all').fillna("")

            if battery_info.shape[1] > 7:
                battery_info = battery_info.drop(columns=battery_info.columns[[6, 7]], errors='ignore')

            battery_info = battery_info.apply(
                lambda column: column.map(
                    lambda value: value.replace(',', '.') if isinstance(value, str) else value
                )
            )

            cols = list(battery_info.columns)
            if "Capacity (new)" in cols and "Nominal Capacity" in cols:
                nominal_capacity = cols.pop(cols.index("Nominal Capacity"))
                cols.insert(cols.index("Capacity (new)") + 1, nominal_capacity)
                battery_info = battery_info[cols]

            for column_name, suffix in {
                "Capacity (new)": " kWh",
                "Nominal Capacity": " Ah",
            }.items():
                if column_name in battery_info.columns:
                    battery_info[column_name] = battery_info[column_name].astype(str).str.strip().replace({'': np.nan})
                    battery_info[column_name] = battery_info[column_name].map(
                        lambda value: f"{value}{suffix}" if isinstance(value, str) and value else value
                    )

            if len(battery_info.columns) > 6:
                column_name = battery_info.columns[6]
                battery_info[column_name] = battery_info[column_name].astype(str).str.strip().replace({'': np.nan})
                battery_info[column_name] = battery_info[column_name].map(
                    lambda value: f"{value} km" if isinstance(value, str) and value else value
                )

            return battery_info.fillna("")

        except Exception as exc:
            st.error(f"Error fetching battery info: {exc}")
            return pd.DataFrame()

    @staticmethod
    def _make_unique_headers(headers: list) -> list:
        """Make column headers unique by adding suffixes to duplicates.

        Args:
            headers: List of column headers.

        Returns:
            List of unique headers.
        """
        unique_header = []
        duplicate_counts = {}

        for column in headers:
            cleaned_column = column.strip()
            if cleaned_column not in unique_header:
                unique_header.append(cleaned_column)
                duplicate_counts[cleaned_column] = 1
            else:
                duplicate_counts[cleaned_column] += 1
                unique_header.append(f"{cleaned_column}_{duplicate_counts[cleaned_column]}")

        return unique_header

    @staticmethod
    def _find_battery_pack_column(df: pd.DataFrame) -> Optional[str]:
        """Find the battery pack column in the DataFrame.

        Args:
            df: DataFrame to search.

        Returns:
            Name of the battery pack column, or None if not found.
        """
        battery_pack_cols = [column for column in df.columns if column.startswith('Battery Pack')]
        return battery_pack_cols[0] if battery_pack_cols else None

    @staticmethod
    def _parse_decimal_series(series: pd.Series, units: Optional[List[str]] = None) -> pd.Series:
        """Parse decimal text data into numeric values."""
        cleaned = series.fillna("").astype(str).str.strip()
        for unit in units or []:
            cleaned = cleaned.str.replace(unit, "", regex=False)
        cleaned = cleaned.replace({'': np.nan, 'nan': np.nan})
        cleaned = cleaned.str.replace(',', '.', regex=False)
        cleaned = cleaned.str.extract(r'(-?\d+(?:\.\d+)?)', expand=False)
        return pd.to_numeric(cleaned, errors='coerce')

    @staticmethod
    def _parse_whole_number_series(series: pd.Series) -> pd.Series:
        """Parse integers that may include thousands separators or units."""
        cleaned = series.fillna("").astype(str).str.strip().replace({'': np.nan, 'nan': np.nan})
        cleaned = cleaned.str.replace(r'[^\d-]', '', regex=True)
        return pd.to_numeric(cleaned, errors='coerce')

    @staticmethod
    def _process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Process and clean the DataFrame.

        Args:
            df: Raw DataFrame from Google Sheets.

        Returns:
            Processed DataFrame.
        """
        processed_df = df.copy()
        battery_pack_col = GoogleSheetsClient._find_battery_pack_column(processed_df)

        object_columns = processed_df.select_dtypes(include='object').columns.tolist()
        if battery_pack_col and battery_pack_col in object_columns:
            object_columns.remove(battery_pack_col)

        for column in object_columns:
            processed_df[column] = processed_df[column].astype(str).str.strip()
            processed_df[column] = processed_df[column].replace({'': np.nan, 'nan': np.nan})

        numeric_decimals = {
            'Age': [' Months'],
            'Capacity Net Now': [' kWh'],
            'Daily SOC Limit': ['%'],
            'DC Ratio': ['%'],
            'Cycles': [],
        }
        for column, units in numeric_decimals.items():
            if column in processed_df.columns:
                processed_df[column] = GoogleSheetsClient._parse_decimal_series(processed_df[column], units)

        if 'Odometer' in processed_df.columns:
            processed_df['Odometer'] = GoogleSheetsClient._parse_whole_number_series(processed_df['Odometer'])

        if 'Rated Range' in processed_df.columns:
            processed_df['Rated Range'] = GoogleSheetsClient._parse_whole_number_series(processed_df['Rated Range'])

        if 'Degradation' in processed_df.columns:
            degradation = GoogleSheetsClient._parse_decimal_series(processed_df['Degradation'], ['%'])
            processed_df['Degradation'] = -degradation.abs()
            processed_df.loc[processed_df['Degradation'] == 0, 'Degradation'] = np.nan
            processed_df['SOH'] = 100 + processed_df['Degradation']

        return processed_df
